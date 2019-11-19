import sys


try:
    import openpyxl as opx
except ImportError as iE:
    print("module not fucking loading")
    sys.exit(1)


def load_workbook(file):
    '''Returns the active excel worksheet'''
    file_name = file
    sheet_name = 'Sheet1'
    wb = opx.load_workbook(file_name)  # loads excel workbook
    ws = wb[sheet_name]  # reads active worksheet into memory
    return wb, ws


def get_data_from_sheet(wb, ws):
    '''Returns a dictionary of column values'''
    column_values = []
    for _ in range(2, 40): # sets range of rows to fetch data from spreadsheet
        info_dict = { # creates a dictionary of name, password, email for each row in spreadsheet
            'name': None,
            'password': None,
            'email': None
        }
        info_dict['name'] = ws.cell(row=_, column=1).value # maps 'name' data to correct column in spreadsheet
        info_dict['password'] = ws.cell(row=_, column=3).value # maps 'password' data to correct column in spreadsheet
        info_dict['email'] = ws.cell(row=_, column=2).value # maps 'email' data to correct column in spreadsheet
        column_values.append(info_dict)
    wb.close()
    return column_values


def write_emails(col_vals):
    '''Reads email.txt and replaces placeholders with values in
    list_of_dicts'''
    email_file = open('email.txt', 'r')
    message_content = email_file.read()
    new_file = open('output.txt', 'w')
    for person in col_vals:
        formatted_message = message_content.format(
            person['name'],
            person['email'],
            person['password'])
        new_file.write(formatted_message+'\n'+'\n')
    email_file.close()
    new_file.close()


def main():
    file = 'example_email_list.xlsx'
    workbook, worksheet = load_workbook(file)
    data = get_data_from_sheet(workbook, worksheet)
    write_emails(data)


if __name__ == '__main__':
    main()
