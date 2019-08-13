import sys

try:
    import openpyxl as opx
except ImportError as iE:
    print("module not fucking loading")
    sys.exit(1)


list_of_dicts = []


def get_name_mail():
    """creates a dictionary of name, email, password and appends to
    list_of_dicts"""
    file_name = 'ENTER/FILE/PATH/example_email_list.xlsx' # input local file path
    sheet_name = 'Sheet1'
    wb = opx.load_workbook(file_name)
    ws = wb[sheet_name]
    for _ in range(2, 40): # sets range of rows to fetch data from spreadsheet
        info_dict = { # creates a dictionary of name, password, email for each row in spreadsheet
            'name': None,
            'password': None,
            'email': None
        }
        info_dict['name'] = ws.cell(row=_, column=1).value # maps 'name' data to correct column in spreadsheet
        info_dict['password'] = ws.cell(row=_, column=3).value # maps 'password' data to correct column in spreadsheet
        info_dict['email'] = ws.cell(row=_, column=2).value # maps 'email' data to correct column in spreadsheet
        list_of_dicts.append(info_dict)
    wb.close()
    return(list_of_dicts)


def write_emails(list_of_dicts):
    """Reads email.txt and replaces placeholders with values in
    list_of_dicts"""
    email_file = open('email.txt', 'r')
    message_content = email_file.read()
    new_file = open('output.txt', 'w')
    for person in list_of_dicts:
        formatted_message = message_content.format(
            person['name'],
            person['email'],
            person['password'])
        new_file.write(formatted_message+'\n'+'\n')
    email_file.close()
    new_file.close()


if __name__ == '__main__':
    get_name_mail()
    write_emails(list_of_dicts)
